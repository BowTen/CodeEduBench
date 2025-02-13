from database_connect import get_session
import aitestOrm
from tabulate import tabulate
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from sqlalchemy.orm import sessionmaker

session = get_session('database_aitest')

# 创建 Excel 文件
wb = Workbook()
ws = wb.active
ws.title = "Matrix"

rd = 1

def gen_matrix(model_name, table_name, sc_table, pai, pman):
    global rd
    # 查询数据
    anno_scs = session.query(sc_table).filter(sc_table.model_name == model_name).all()
    
    # 初始化一个 7x7 的零矩阵
    matrix = [[0 for _ in range(7)] for _ in range(7)]
    
    # 遍历查询结果，更新矩阵
    for anno_sc in anno_scs:
        row = getattr(anno_sc, pai)  # 获取"pai"对应的列的值
        col = getattr(anno_sc, pman)  # 获取"pman"对应的列的值
        if row > 5:
            row = 5
        if col > 5:
            col = 5
        matrix[row][col] += 1  # 增加计数

    # 计算“总计”行和列
    for i in range(6):
        matrix[i][6] = sum(matrix[i][:6])  # 计算每行的总计
        matrix[6][i] = sum(matrix[j][i] for j in range(6))  # 计算每列的总计
    matrix[6][6] = sum(matrix[6][:6])  # 计算总计的总计
    
    for i in range(7):
        for j in range(7):
            print(matrix[i][j], end=' ')
        print()

    ws.cell(row=rd, column=1, value=model_name + "模型  " + table_name + "  " + pai)  # 设置表格标题
    # 填充数据
    for i in range(7):
        for j in range(7):
            ws.cell(row=rd+i+1, column=j+1, value=matrix[i][j])
    
    # 设置表头的斜线
    header = ws.cell(row=rd+1, column=1)
    header.value = "人工" + model_name
    
    # 使用斜线分割
    header.alignment = Alignment(horizontal='center', vertical='center')
    header.border = Border(top=Side(border_style='thin'), right=Side(border_style='thin'))
    
    # 设置第一行和第一列标题
    for i in range(1, 7):
        ws.cell(row=rd+1, column=i+1, value=str(i))
        ws.cell(row=rd+i+1, column=1, value=str(i))
    
    ws.cell(row=rd+1, column=7, value="总计")
    ws.cell(row=rd+7, column=1, value="总计")
    
    rd += 9  # 移动到下一个表格的起始行

    
if __name__ == '__main__':
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Annotation Score", aitestOrm.AnnotationGen, "accuracy", "maccuracy")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Annotation Score", aitestOrm.AnnotationGen, "simplicity", "msimplicity")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Annotation Score", aitestOrm.AnnotationGen, "naturalness", "mnaturalness")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Annotation Score", aitestOrm.AnnotationGen, "usefulness", "musefulness")
    
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Knowledge generate Score", aitestOrm.KnlgExp, "accuracy", "maccuracy")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Knowledge generate Score", aitestOrm.KnlgExp, "correlation", "mcorrelation")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Knowledge generate Score", aitestOrm.KnlgExp, "understandability", "munderstandability")
    
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Test Case generate Score", aitestOrm.CaseGen, "comprehensive", "mcomprehensive")

    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Code generate Score", aitestOrm.CodeGen, "readability", "mreadability")
    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Code generate Score", aitestOrm.CodeGen, "performance", "mperformance")

    gen_matrix("Qwen2.5-Coder-7B-Instruct", "Code correcting Score", aitestOrm.CodeCor, "understandability", "munderstandability")
    
    
    
    wb.save("output_matrix_with_slash.xlsx")